import io
import json
from pathlib import Path
import zipfile

import streamlit as st
from openpyxl import load_workbook

from transform_sections import process_workbook_json


BASE_DIR = Path(__file__).resolve().parent
TMP_DIR = BASE_DIR / "tmp"
TMP_DIR.mkdir(exist_ok=True)


def excel_to_workbook_dict(file_bytes: bytes, file_name: str) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets: dict[str, dict] = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        segments: list[tuple[int, int]] = []
        in_segment = False
        start_idx = 0

        for idx, row in enumerate(rows):
            has_data = any(
                (cell is not None and str(cell).strip() != "") for cell in row
            )
            if has_data and not in_segment:
                in_segment = True
                start_idx = idx
            elif not has_data and in_segment:
                segments.append((start_idx, idx - 1))
                in_segment = False

        if in_segment:
            segments.append((start_idx, len(rows) - 1))

        tables = []
        for start, end in segments:
            segment_rows = rows[start : end + 1]

            min_col = None
            max_col = None
            for r in segment_rows:
                for j, cell in enumerate(r):
                    if cell is not None and str(cell).strip() != "":
                        if min_col is None or j < min_col:
                            min_col = j
                        if max_col is None or j > max_col:
                            max_col = j

            if min_col is None:
                continue

            data = []
            for r in segment_rows:
                row_vals = []
                for j in range(min_col, max_col + 1):
                    if j < len(r):
                        row_vals.append(r[j])
                    else:
                        row_vals.append(None)
                data.append(row_vals)

            tables.append(
                {
                    "start_row": start + 1,
                    "start_col": min_col + 1,
                    "row_count": len(data),
                    "column_count": max_col - min_col + 1,
                    "data": data,
                }
            )

        sheets[ws.title] = {"tables": tables}

    return {"file_name": file_name, "sheets": sheets}


def transform_uploaded_file(file_bytes: bytes, file_name: str) -> dict | None:
    suffix = Path(file_name).suffix.lower()

    if suffix == ".json":
        tmp_path = TMP_DIR / f"raw_{file_name}"
        tmp_path.write_bytes(file_bytes)
        return process_workbook_json(tmp_path)

    if suffix in (".xlsx", ".xls"):
        workbook_dict = excel_to_workbook_dict(file_bytes, file_name)
        tmp_json_path = TMP_DIR / f"{Path(file_name).stem}_workbook.json"
        tmp_json_path.write_text(
            json.dumps(workbook_dict, ensure_ascii=False), encoding="utf-8"
        )
        return process_workbook_json(tmp_json_path)

    st.error("Unsupported file type. Please upload .json, .xlsx or .xls.")
    return None


def apply_theme(theme: str) -> None:
    if theme == "Dark":
        css = """
        <style>
        body, .main {
            background-color: #0e1117;
            color: #e5e7eb;
        }
        section.main > div {
            background-color: #0e1117;
        }
        pre, code {
            background-color: #111827 !important;
            color: #e5e7eb !important;
        }
        .stDownloadButton button {
            background-color: #1f2937;
            color: #e5e7eb;
        }
        </style>
        """
    else:
        css = """
        <style>
        pre, code {
            background-color: #f3f4f6 !important;
            color: #111827 !important;
        }
        </style>
        """
    st.markdown(css, unsafe_allow_html=True)


def main():
    st.set_page_config(page_title="GST Workbook Transformer", layout="wide")

    left, right = st.columns([3, 1])
    with left:
        st.title("GST Excel → Structured JSON")
        st.write(
            "Upload GST workbook as Excel (.xlsx / .xls) or as raw workbook JSON. "
            "The app converts it into clean structured JSON."
        )
    with right:
        theme = st.radio("Theme", ["Light", "Dark"], index=0, horizontal=True)
    apply_theme(theme)

    uploaded_files = st.file_uploader(
        "Upload Excel or workbook JSON file(s)",
        type=["json", "xlsx", "xls"],
        accept_multiple_files=True,
    )

    if not uploaded_files:
        return

    results: list[tuple[str, dict]] = []

    for upl in uploaded_files:
        file_bytes = upl.read()
        try:
            structured = transform_uploaded_file(file_bytes, upl.name)
        except Exception as e:
            st.error(f"Error while processing {upl.name}: {e}")
            continue

        if structured:
            results.append((upl.name, structured))
        else:
            st.warning(f"No structured tables produced for {upl.name}.")

    if not results:
        return

    st.markdown("---")
    st.subheader("Structured JSON Preview")

    max_cols = 3
    for i in range(0, len(results), max_cols):
        row = results[i : i + max_cols]
        cols = st.columns(len(row))
        for col, (name, data) in zip(cols, row):
            with col:
                st.markdown(f"**{name}**")
                st.success("Structured JSON generated")
                formatted = json.dumps(data, ensure_ascii=False, indent=2)
                st.code(formatted, language="json")
                out_name = f"structured_{Path(name).stem}.json"
                st.download_button(
                    label="Download JSON",
                    file_name=out_name,
                    mime="application/json",
                    data=formatted.encode("utf-8"),
                    key=f"dl_json_{out_name}",
                )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in results:
            fname = f"structured_{Path(name).stem}.json"
            zf.writestr(fname, json.dumps(data, ensure_ascii=False, indent=2))
    zip_buffer.seek(0)

    st.markdown("---")
    st.download_button(
        "Download all structured JSON as ZIP",
        data=zip_buffer.getvalue(),
        file_name="structured_json_bundle.zip",
        mime="application/zip",
        key="dl_zip_all",
    )


if __name__ == "__main__":
    main()
